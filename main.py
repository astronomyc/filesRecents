# Importaciones Requeridas
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Obtener los archivos mas recientes de los directorios
def get_most_recents_file(rute):
    recent_file = None
    recent_date = None
    for directory, subdirectories, files in os.walk(rute):
        for file in files:
            file_rute = os.path.join(directory, file)
            file_date = os.path.getmtime(file_rute)
            if recent_file is None or file_date > recent_date:
                recent_file = file
                recent_date = file_date
    return recent_file, recent_date

# Formatear tiempo a fecha DD/MM/YYYY
def format_date(timestamp):
    return datetime.datetime.fromtimestamp(timestamp).strftime('%d/%m/%Y')

# Calcular peso de un directorio
def get_folder_size(folder):
    total_size = 0
    for dirpath, dirnames, filenames in os.walk(folder):
        for file in filenames:
            file_path = os.path.join(dirpath, file)
            total_size += os.path.getsize(file_path)
    return total_size

# Convertir el formato de bytes legible para los humanos
def convert_bytes_to_human_readable(size_in_bytes):
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if size_in_bytes < 1024:
            return f"{size_in_bytes:.2f} {unit}"
        size_in_bytes /= 1024

# Solicitar ruta al usuario
rute_dir = input("Ruta: ").strip()

# Iniciar libro de Excel
wb = Workbook()
ws = wb.active

# Configurar ancho de las columnas
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 48
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12

# Escribir las primeras dos filas
ws.append(["Archivos Recientes", "", "", ""])
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws.append(["Carpeta", "Nombre del archivo", "Fecha", "Peso"])

for col in range(1, 5):
    ws.cell(row=1, column=col).fill = PatternFill(start_color="BB88FF", end_color="BB88FF", fill_type="solid")
    ws.cell(row=2, column=col).fill = PatternFill(start_color="EECCFF", end_color="EECCFF", fill_type="solid")
    ws.cell(row=2, column=col).font = Font(bold=True)

for folder in os.listdir(rute_dir):
    folder_path = os.path.join(rute_dir, folder)
    if os.path.isdir(folder_path):
        # Obtener el archivo más reciente, la fecha y el tamaño de la carpeta
        file_recent, date_recent = get_most_recents_file(folder_path)
        formatted_date = format_date(date_recent)
        folder_name = os.path.basename(folder_path)
        folder_size = get_folder_size(folder_path)

        # Cambiar Bytes a Legible
        formatted_size = convert_bytes_to_human_readable(folder_size)

        # Escribir los datos en el archivo Excel
        ws.append([folder_name, file_recent, formatted_date, formatted_size])

# Guardar excel
wb.save(os.path.join(rute_dir, f"ArchivosRecientes.xlsx"))
print('Complete')